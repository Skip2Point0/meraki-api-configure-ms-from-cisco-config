import requests
import json
import re
from openpyxl import load_workbook
# ################################################# PARAMETERS #########################################################
# Configure parameters below: Meraki APIkey, Organizaton Name, Networks where switches will be installed. Spreadsheet
# name and Physical Addresses of destination networks. Address is only necessary if devices will be claimed and
# configured with this script.
# ######################################################################################################################
meraki_api = "MerakiAPIkey"
organization_id = 'Organization Name'
network_ids = ['Site Name1', 'Site Name2']
spreadsheet = 'Switch Spreadsheet.xlsx'
address = ['123 ABC Street, Somewhere USA', '456 DEF Street, Somewhere USA']
# sws_number = '1'
# ######################################################################################################################
# Configurations below this line are optional. If you are using your own spreadsheet, change column letters below
# ######################################################################################################################
ms_name_column = 'C'
cisco_ipcolumn = 'B'
notes_column = 'D'
tags_column = 'E'
serials_column = 'F'
mac_column = 'G'
# ######################################################################################################################
interface_list = {}
shard_url = ()
configfiles = []
serials = []
device_names = []
net_dictionary = {}
headers = {
    'X-Cisco-Meraki-API-Key': meraki_api,
    'Content-Type': 'application/json'
    }
wb = load_workbook(spreadsheet)


def calculate_interface(cnf):
    interfaces = []
    interface_line = []
    for num_lines in range(len(cnf)):
        line = cnf[num_lines]
        reg = re.compile(r'interface\s(TenGigabitEthernet|GigabitEthernet|FastEthernet|Ethernet)(\d.\d.\d+)')
        interface = reg.search(line)
        reg = re.compile(r'interface\sPort-channel(\d+)')
        port_channel = reg.search(line)
        if interface:
            interfaces.append(interface.group(1) + interface.group(2))
            interface_line.append(num_lines)
        elif port_channel:
            po = "Po" + port_channel.group(1)
            interfaces.append(po)
            interface_line.append(num_lines)
        else:
            continue
    return interfaces, interface_line


def calculate_interface_config(cnf, int_line, inter):
    interface_dictionary = []
    for index in range(len(int_line)):
        interface_structure = {"number": [],
                               "name": [],
                               "enabled": [],
                               "type": [],
                               "vlan": [],
                               "voiceVlan": [],
                               "allowedVlans": [],
                               "rstpEnabled": [],
                               "stpGuard": [],
                               "portChannel": [],
                               "portChannelMode": [],
                               "linkNegotiation": []}
        object_line = int_line[index]
        for line in cnf[object_line + 1:]:
            if line != "!":
                reg = re.compile(r'\sswitchport\saccess\svlan\s(\d+)')
                v = reg.search(line)
                reg = re.compile(r'\sswitchport\svoice\svlan\s(\d+)')
                vv = reg.search(line)
                reg = re.compile(r'\sswitchport\strunk\snative\svlan\s(.*)')
                nv = reg.search(line)
                reg = re.compile(r'\sswitchport\strunk\sallowed\svlan\s(.+)')
                av = reg.search(line)
                reg = re.compile(r'\sswitchport\smode\s(access|trunk)')
                pm = reg.search(line)
                reg = re.compile(r'\sdescription\s(.*)')
                descr = reg.search(line)
                reg = re.compile(r'\s(shut)')
                shut = reg.search(line)
                reg = re.compile(r'\sspanning-tree\s(portfast)')
                portf = reg.search(line)
                reg = re.compile(r'\sspanning-tree\sbpduguard\s(\w+)')
                bpdug = reg.search(line)
                reg = re.compile(r'\schannel-group\s(\d+)\smode\s(\w+)')
                chgrp = reg.search(line)
                reg = re.compile(r'\snegotiation\s(\w+)')
                portsp = reg.search(line)
                interface_structure["number"] = inter[index]
                if descr:
                    # print(descr.group(1))
                    interface_structure["name"] = descr.group(1)
                elif pm:
                    # print(pm.group(1))
                    interface_structure["type"] = pm.group(1)
                elif v:
                    # print(v.group(1))
                    interface_structure["vlan"] = v.group(1)
                elif nv:
                    interface_structure["vlan"] = nv.group(1)
                elif vv:
                    # print(av.group(1))
                    interface_structure["voiceVlan"] = vv.group(1)
                elif av:
                    interface_structure["allowedVlans"] = av.group(1)
                elif shut:
                    interface_structure["enabled"] = "false"
                elif portf:
                    interface_structure["rstpEnabled"] = "true"
                elif bpdug:
                    interface_structure["stpGuard"] = "true"
                elif chgrp:
                    interface_structure["portChannel"] = chgrp.group(1)
                    interface_structure["portChannelMode"] = chgrp.group(2)
                elif portsp:
                    interface_structure["linkNegotiation"] = portf.group(1)
            else:
                interface_structure["number"] = inter[index]
                interface_dictionary.append(interface_structure)
                break
    return interface_dictionary


def parse_switch_config():
    load_from_spreadsheet()
    ctr = 0
    print(configfiles)
    for f in configfiles:
        with open(f) as g:
            config = g.read().splitlines()
        reg = re.compile(r'(\d+.\d+.\d+.\d+)_show_run.txt')
        switch_ip = reg.search(f)
        switch_ip = switch_ip[1]
        interface = calculate_interface(config)
        calculated = calculate_interface_config(config, interface[1], interface[0])
        interface_list[switch_ip] = calculated
        ctr = ctr + 1
    print(interface_list)


def meraki_port_structure(i):
    reg = re.compile(r'(TenGigabit|Gigabit|Fast)Ethernet\d.(\d).(\d+)')
    number = reg.search(i['number'])
    if number and number[2] == '0':
        if number[1] == 'Gigabit' or 'Fast':
            if i['name']:
                name = i['name']
            else:
                name = None
            if i['enabled']:
                enabled = i['disabled']
            else:
                enabled = 'true'
            if i['type']:
                swtype = i['type']
            else:
                swtype = 'access'
            if i['vlan']:
                vlan = i['vlan']
            else:
                vlan = '1'
            if i['voiceVlan']:
                voicevlan = i['voiceVlan']
            else:
                voicevlan = None
            if i['allowedVlans'] and i['type'] == 'trunk':
                allowedvlans = i['allowedVlans']
            else:
                allowedvlans = 'all'
            if i['rstpEnabled']:
                rstp = i['rstpEnabled']
            else:
                rstp = 'false'
            if i['stpGuard']:
                stpguard = i['stpGuard']
            else:
                stpguard = 'disabled'
            payload = {
                "name": name,
                "enabled": enabled,
                "type": swtype,
                "vlan": vlan,
                "voiceVlan": voicevlan,
                "allowedVlans": allowedvlans,
                "rstpEnabled": rstp,
                "stpGuard": stpguard,
                # "linkNegotiation": link
            }
            payload = json.dumps(payload)
            return payload


def pull_organization_id(head):
    global shard_url
    url = "https://api.meraki.com/api/v0/organizations"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    response = json.loads(response)
    for dicti in response:
        name = dicti["name"]
        if name == organization_id:
            org_id = dicti["id"]
            shard_url = dicti["url"]
            urllenght = shard_url.find('com') + 3
            shard_url = shard_url[:urllenght]
            print("\n#################################################\n")
            print(name + "\n" + "Organization ID: " + org_id)
            print("Organization Shard URL: " + shard_url)
            print("\n#################################################\n")
            return org_id
        else:
            continue


def pull_organization_networks(head):
    global net_dictionary
    global organization_id
    organization_id = pull_organization_id(head)
    url = shard_url + "/api/v0/organizations/" + organization_id + "/networks"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    json_response = json.loads(response)
    # print(response)
    for networks in json_response:
        name = networks['name']
        n_id = networks['id']
        # print(name + " : " + n_id)
        net_dictionary[name] = n_id
    # print(net_dictionary)
    return net_dictionary


def pull_destination_networks():
    global network_ids
    dest_network_ids = []
    for n in network_ids:
        for i in net_dictionary:
            if n == i:
                print("Destination Network: " + n)
                dest_network_ids.append(net_dictionary[n])
                break
            else:
                continue
    print(dest_network_ids)
    return dest_network_ids


pull_organization_networks(headers)
networks_dest = pull_destination_networks()


def meraki_claim_serial(network, addr, head):
    incr = 0
    for net in network:
        print("\n########################################")
        print("NETWORK:  " + net)
        for sl in serials:
            print("########################################")
            print("CLAIMING SERIAL NUMBER - " + sl)
            print("########################################")
            url = shard_url + "/api/v0/networks/" + net + "/devices/claim"
            print(url)
            payload = {
                "serial": sl,
            }
            payload = json.dumps(payload)
            print(payload)
            response = requests.request("POST", url, headers=head, data=payload)
            print(response.text.encode('utf8'))

            url = shard_url + "/api/v0/networks/" + net + "/devices/" + sl
            print(url)
            payload = {
                "name": device_names[incr],
                "address": addr[incr]
            }
            payload = json.dumps(payload)
            print(payload)
            response = requests.request("PUT", url, headers=head, data=payload)
            print(response.text.encode('utf8'))
        incr = incr + 1


def load_from_spreadsheet():
    for sheet in network_ids:
        sheet = wb[sheet]
        incr = 2
        maximum = sheet.max_row - 1
        for row in range(maximum):
            a = sheet[cisco_ipcolumn + str(incr)].value
            s = sheet[serials_column + str(incr)].value
            n = sheet[ms_name_column + str(incr)].value
            configfiles.append(str(a) + '_show_run.txt')
            serials.append(str(s))
            device_names.append((str(n)))
            incr = incr + 1


def meraki_ms_config(head):
    global interface_list
    parse_switch_config()
    incr = 0
    for i in interface_list:
        print("\n########################################\n")
        print("APPLYING MS CONFIG... " + i + " >>> " + device_names[incr] + "\n")
        sl = serials[incr]
        incr = incr + 1
        for interfaces in (interface_list[i]):
            reg = re.compile(r'(TenGigabit|Gigabit|Fast)Ethernet(\d).(\d).(\d+)')
            number = reg.search(interfaces['number'])
            # if number and number[2] == swsn:
            url = shard_url + "/api/v0/devices/" + sl + "/switchPorts/" + number[4]
            print(url)
            payload = meraki_port_structure(interfaces)
            print(payload)
            response = requests.request("PUT", url, headers=head, data=payload)
            f = response.content
            print(f)


load_from_spreadsheet()
# ################################################# PARAMETERS #########################################################
# meraki_claim_serial(networks_dest, address, headers)
meraki_ms_config(headers)
# ######################################################################################################################
